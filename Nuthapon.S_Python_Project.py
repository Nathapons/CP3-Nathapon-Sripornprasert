import openpyxl as xl
from openpyxl.chart import ScatterChart, Reference, Series
from tkinter import *
from tkinter import ttk
from tkinter import messagebox as msgbox
from tkinter import filedialog as fd
import os


def program_overview():
    need_to_import_program = True
    file_type = str(save_file_type.get())

    if file_type.startswith("Pdf"):
        alarm_user = msgbox.showwarning(title="Warining", message="Program is not ready")
    elif file_type.startswith("Excel"):

        while need_to_import_program:
            # Run function
            format_book, result_book, result_file = open_excel_file()
            fill_result_information(format_book, result_book)
            create_worksheet(format_book, result_book, result_file)
            save_format_file(format_book, result_book)

            # Loop command
            need_to_import_program = msgbox.askyesno(title="Ask User", message="คุณต้องการจะอิมพอทเอกสารต่อหรือไม่?")

        # if user don't operate
        quit()

    else:
        alarm_user = msgbox.showwarning(title="Warining", message="กรุณากรอก Save as file type ด้วย")
        # if user don't operate
        quit()


def open_excel_file():
    # Valiable Assignment
    initial_dir = os.path.abspath(__file__)

    # User select file
    information_message = msgbox.showinfo(title="Information", message="กรุณาเลือกไฟล์ Format.xlsx")
    format_file = fd.askopenfilename(initialdir=initial_dir, title="กรุณาเลือกไฟล์ Format.xlsx")
    information_message = msgbox.showinfo(title="Information", message="กรุณาเลือกไฟล์ Result.xlsx")
    result_file = fd.askopenfilename(initialdir=initial_dir, title="กรุณาเลือกไฟล์ Result.xlsx")

    # Get Load workbook
    try:
        # User select 2 file
        format_book = xl.load_workbook(filename=format_file)
        result_book = xl.load_workbook(filename=result_file)

        # Run function: create_worksheet
        return format_book, result_book, result_file

    except:
        # User don't select 2 file
        alarm_user = msgbox.showwarning(title="Warining", message="การเลือกไฟล์ได้ถูกยกเลิก")
        quit()


def fill_result_information(format_book, result_book):
    # กำหนดค่าตัวแปร
    format_sheet = format_book['1']
    tensile_condition_sheet = result_book['UniTest.TensileCond']
    row_in_result_sheet = 1

    while row_in_result_sheet <= tensile_condition_sheet.max_row:
        # ทำการสร้างเซลที่พร้อมสำหรับวางข้อมูลก่อน
        condition_name = tensile_condition_sheet.cell(row=row_in_result_sheet, column=2).value
        condition_value = tensile_condition_sheet.cell(row=row_in_result_sheet, column=3).value
        condition_unit = tensile_condition_sheet.cell(row=row_in_result_sheet, column=4).value
        if str(condition_unit).startswith("None"):
            condition_unit = ""
        condition_result = str(condition_value) + " " + str(condition_unit)

        cell_paste_dict1 = {"Sample name": {"cell1": "H4", "cell2": "G28"},
                            "Lot No.": {"cell1": "H5", "cell2": "B29"},
                            "Operator": {"cell1": "H7", "cell2": "G29"}}
        cell_paste_dict2 ={"Test speed": "G24", "Test date": "B27", "Humidity": "B28",
                           "Comment1": "B30", "Temperature": "G27", "Comment2": "G30"}

        # ทำการ Import information ลงในแบบฟอร์ม
        # กรณีที่วาง 2 cell
        if condition_name in cell_paste_dict1:
            range1 = cell_paste_dict1[condition_name]["cell1"]
            range2 = cell_paste_dict1[condition_name]["cell2"]
            format_sheet[range1] = condition_result
            format_sheet[range2] = condition_result
        # กรณีที่วาง 1 cell
        elif condition_name in cell_paste_dict2:
            range1 = cell_paste_dict2[condition_name]
            format_sheet[range1] = condition_result
        # กรณีที่ข้อมูลเป็น Load cell rating
        elif str(condition_name).startswith("Load cell rating"):
            format_sheet['B24'] = condition_result
            format_sheet['G23'] = str(condition_value / 100)

        # Loop command
        row_in_result_sheet += 1


def create_worksheet(format_book, result_book, result_file):
    # กำหนดค่าตัวแปร
    data_sheet = result_book['Data']
    current_format_sheet = format_book.active
    last_format_sheet_name = data_sheet.cell(row=data_sheet.max_row, column=1).value

    for row_in_data_sheet in range(4, data_sheet.max_row + 1, 1):
        # Get current_format_sheet name
        test_number = data_sheet.cell(row=row_in_data_sheet, column=1).value

        # Run function
        fill_data_result(format_book, current_format_sheet, data_sheet)
        sheet_name_active = decision_data_sheet(test_number)
        scatter_chart = create_graph(result_file, result_book, sheet_name_active, current_format_sheet)
        copy_and_paste_graph(current_format_sheet, result_book, sheet_name_active, scatter_chart)

        # Create_worksheet
        if test_number < last_format_sheet_name:
            next_test_number = data_sheet.cell(row=row_in_data_sheet + 1, column=1).value
            new_format_sheet = format_book.copy_worksheet(current_format_sheet)
            new_format_sheet.title = str(next_test_number)
            current_format_sheet = format_book[str(next_test_number)]


def fill_data_result(format_book, current_format_sheet, data_sheet):
    # Valiable Assignment
    column_in_data_and_format_sheet = 2

    # Fill test number number
    data_sheet['A35'] = current_format_sheet.title

    # Fill test number result and unit
    while column_in_data_and_format_sheet <= data_sheet.max_column:
        test_number_unit = data_sheet.cell(row=3, column=column_in_data_and_format_sheet).value
        current_format_sheet.cell(row=34, column=column_in_data_and_format_sheet).value = test_number_unit
        # Loop command
        column_in_data_and_format_sheet += 1


def decision_data_sheet(test_number):
    # Get name will plot graph
    if test_number >= 100:
        sheet_name_active = "SS-Carve" + str(test_number)
    elif test_number >= 10:
        sheet_name_active = "SS-Carve0" + str(test_number)
    else:
        sheet_name_active = "SS-Carve00" + str(test_number)

    # Return value instead of function
    return sheet_name_active


def create_graph(result_file, result_book, sheet_name_active, current_format_sheet):
    # Valiable Assignment
    book_name_will_plot_graph = os.path.basename(result_file)
    sheet_name_will_plot_graph = result_book[sheet_name_active]
    max_row_in_sheet_name = sheet_name_will_plot_graph.max_row

    # Assign value in X-axis and Y-axis
    scatter_chart = ScatterChart()
    y_values = Reference(sheet_name_will_plot_graph, min_row=3, max_row=max_row_in_sheet_name, min_col=3)
    x_values = Reference(sheet_name_will_plot_graph, min_row=3, max_row=max_row_in_sheet_name, min_col=2)
    series = Series(y_values, x_values, title=None, title_from_data=False)
    scatter_chart.series.append(series)

    # Adjust Scatter Element
    scatter_chart.y_axis.scaling.min = 0
    scatter_chart.legend = None
    scatter_chart.y_axis.number_format = "0.00"

    # Paste Scatter_chart
    current_format_sheet.add_chart(scatter_chart, "A5")

    # Setting width and height of scatter_chart
    scatter_chart.width = 14
    scatter_chart.height = 8
    return scatter_chart


def copy_and_paste_graph(current_format_sheet, result_book, sheet_name_active, scatter_chart):
    pass


def save_format_file(format_book, result_book):
    result_book.save("Y.xlsx")
    new_format_book_name = "X"
    format_book.save(new_format_book_name + ".xlsx")


# Create Window interface
main_window = Tk()
main_window.geometry("500x180")
main_window.title("Peel Stength Program")

# Input Widget
frame1 = Frame(main_window)
welcome_label = ttk.Label(main_window, text="Peel Strength Report Program",
                          font=("Monospaced", 24), foreground="white", background="blue")
save_as_label = ttk.Label(frame1, text="Save as file type:", font=("Arial", 18))
save_file_type = ttk.Combobox(frame1, values=["Excel file", "Pdf file"], font=("Arial", 18),
                              width=15, justify="center")
next_button = Button(frame1, text="Next Program", command=program_overview, font=("Arial", 18))
close_button = Button(frame1, text="Close Program", command=quit, font=("Arial", 18))

# Place widget in Window interface
frame1.place(x=50, y=50)
welcome_label.place(x=40, y=10)
save_as_label.grid(row=1, column=0, ipady=10)
save_file_type.grid(row=1, column=1)
next_button.grid(row=2, column=1, ipadx=5)
close_button.grid(row=2, column=0, ipadx=5)
main_window.mainloop()